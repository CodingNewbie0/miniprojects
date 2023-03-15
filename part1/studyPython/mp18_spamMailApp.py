# 대량 메일 전송
# pip install openpyxl
from openpyxl import load_workbook 
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


wbook = load_workbook('./studyPython/SpamMailList.xlsx', data_only=True)
wsheet = wbook.active # sheet1 선택

for i in range(1, wsheet.max_row+1):
    recv_email = wsheet.cell(i, 1).value
    print(recv_email)
    try:
        # 실제 메일전송 로직
        send_mail = 'XXXXXXX@naver.com' # 실제 아이디
        send_pass = 'XXXXXXXX' # 실제 비밀번호
        smtp_name = 'smtp.naver.com'
        smtp_port = 587
        msg = MIMEMultipart()
        msg['Subject'] = '엑셀에서 보내는 메일!!'
        msg['From'] = send_mail
        msg['To'] = send_pass
        msg.attach(MIMEText('부대입니다. 연락주세요.')) # 메일 내용 쓰기

        mail = smtplib.SMTP(smtp_name, smtp_port) # 객체생성
        mail.starttls() # 보안
        mail.login(send_mail, send_pass) # 로그인
        mail.sendmail(send_mail, recv_email, msg.as_string())
        mail.quit()
        print(f'전송성공 : {recv_email}')

    except Exception as e:
        print(f'수신메일 - {recv_email}')
        print(f'전송에러 : {e}')